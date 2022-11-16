from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = {
	"GGG": {
		"strenght": 99,
		"legnth": 78,
		"power": 94,
		"stamina": 89
	},
	"Bivol": {
		"strenght": 79,
		"legnth": 80,
		"power": 75,
		"stamina": 99
	},
	"AJ": {
		"strenght": 90,
		"legnth": 95,
		"power": 93,
		"stamina": 77
	},
	"Fury": {
		"strenght": 99,
		"legnth": 99,
		"power": 80,
		"stamina": 99
	},
	"Inoue": {
		"strenght": 90,
		"legnth": 82,
		"power": 99,
		"stamina": 97
	}
}

wb = Workbook()
ws = wb.active
ws.title = "Grades"

headings = ['Name'] + list(data['GGG'].keys())
ws.append(headings)

for person in data:
	grades = list(data[person].values())
	ws.append([person] + grades)

for col in range(2, len(data['GGG']) + 2):
	char = get_column_letter(col)
	ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

for col in range(1, 6):
	ws[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF")

wb.save("NewRankings.xlsx")